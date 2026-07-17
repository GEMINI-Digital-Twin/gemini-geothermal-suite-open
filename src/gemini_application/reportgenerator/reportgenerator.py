"""PDF report generation for well performance analysis and monitoring data."""

import io
import json
import os
import re
from datetime import datetime, timezone
from math import ceil, sqrt
from pathlib import Path

import matplotlib
import matplotlib.dates as mdates
import matplotlib.image as mpimg
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import numpy as np
import pandas as pd
from matplotlib import rcParams
from matplotlib.backends.backend_pdf import PdfPages
from matplotlib.offsetbox import AnnotationBbox, OffsetImage
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string

from gemini_application.application_abstract import ApplicationAbstract
from gemini_application.injectionwell.injectionwell_monitoring import InjectionWellMonitoring
from gemini_model.fluid.pvt_water_stp import PVTConstantSTP
from gemini_model.reservoir.reservoir_pressuredrop import bottomhole_skin_dp
from gemini_model.well.pressure_drop import DPDT

matplotlib.use("Agg")


class ReportGenerator(ApplicationAbstract):
    """Class for generating reports.

    The class retrieves data from the database and generates a report in pdf format.
    """

    def __init__(self):
        """Initialize report generator."""
        super().__init__()
        self.plant_name = None
        self.project_path = None
        self.start_time = None
        self.end_time = None
        self.start_datestamp_title = None
        self.end_datestamp_title = None
        self.timestep = 3600  # Default value is 1 hour since usually values are given /h
        self.database_internal = None
        self.database_external = None
        self.pdf_buffer = None
        self.pdf_object = None
        self.author_name = None
        self.project_name = None
        self.number_days = None
        self.page_size = (11.69, 8.27)

        # Well pressure drop model
        self.well_DP = DPDT()
        self.well_DP.PVT = PVTConstantSTP()
        self.bottomhole_skin_dp = bottomhole_skin_dp()

    def init_parameters(self, **kwargs):
        """Initialize parameters."""
        for key, value in kwargs.items():
            setattr(self, key, value)

        start_dt = datetime.strptime(self.start_time, "%Y-%m-%d %H:%M:%S")
        self.start_datestamp_title = (
            f"{start_dt.month}/"
            f"{start_dt.day}/"
            f"{start_dt.year} "
            f"{start_dt.strftime('%H')}:{start_dt.strftime('%M')}"
        )

        end_dt = datetime.strptime(self.end_time, "%Y-%m-%d %H:%M:%S")
        self.end_datestamp_title = (
            f"{end_dt.month}/"
            f"{end_dt.day}/"
            f"{end_dt.year} "
            f"{end_dt.strftime('%H')}:{end_dt.strftime('%M')}"
        )

    def calculate(self):
        """Calculate report data."""
        # Class ReportGenerator does not require calculations
        pass

    def get_units(self, tagname):
        """Get units for given tagname."""
        tag = tagname.lower()
        if "pressure" in tag:
            units = "[bar]"
        elif "temperature" in tag:
            units = "[°C]"
        elif "flow" in tag:
            units = "[m^3/h]"
        elif "frequency" in tag:
            units = "[Hz]"
        elif "current" in tag:
            units = "[A]"
        elif "power" in tag:
            units = "[kW]"
        else:
            units = "[-]"
        return units

    def get_data(self, tagname):
        """Get data for given tagname."""
        if tagname is None or tagname == "":
            return None, None  # return both result and time as None

        result, time = self.plant.database.read_internal_database(
            self.unit.plant.name,
            self.unit.name,
            tagname,
            self.start_time,
            self.end_time,
            self.timestep,
        )
        return result, time

    def initialize_pdf_object(self):
        """Initialize PDF object."""
        self.pdf_buffer = io.BytesIO()
        self.pdf_object = PdfPages(self.pdf_buffer)
        return

    def add_title_page(self):
        """Add title page to PDF."""
        title = f"{self.project_name} Report"
        date_str = datetime.now().strftime("%A, %B %dth %Y, %I:%M %p")
        author = self.author_name

        if not hasattr(self, "pdf_object") or self.pdf_object is None:
            raise ValueError(
                "PDF object is not initialized. " "Ensure self.pdf_object is properly set."
            )

        fig, ax = plt.subplots()
        ax.axis("off")  # Remove axes

        # Title: centered in the page
        ax.text(0.5, 0.5, title, fontsize=32, fontweight="bold", ha="center", va="center")

        # Author and Date: below the title, left aligned
        ax.text(0.1, 0.42, f"Date: {date_str}", fontsize=14, ha="left", va="top")
        ax.text(0.1, 0.38, f"Author: {author}", fontsize=14, ha="left", va="top")

        # Logo at top-right
        try:
            logo_img = mpimg.imread(".\\static\\images\\gemini_DDT_V1_300dpi.jpg")
            imagebox = OffsetImage(logo_img, zoom=0.2)
            ab = AnnotationBbox(imagebox, (0.95, 0.92), frameon=False, box_alignment=(1, 1))
            ax.add_artist(ab)
        except FileNotFoundError:
            # Placeholder if no logo image
            ax.text(0.95, 0.92, "[Logo Here]", fontsize=12, ha="right", va="top", style="italic")

        # Save to PDF
        width, height = 11.69, 8.27
        fig.set_size_inches(width, height)
        self.pdf_object.savefig(fig, bbox_inches="tight", pad_inches=0.5)
        plt.close(fig)

    def get_injection_wells(self):
        """Get injection wells component names."""
        output = list()
        for unit in self.plant.units:
            unit_name = unit.name
            if "injection_well" in unit_name:
                output.append(unit_name)
        return output

    def get_production_wells(self):
        """Get production wells component names."""
        output = list()
        for unit in self.plant.units:
            unit_name = unit.name
            if "production_well" in unit_name:
                output.append(unit_name)
        return output

    def get_esps(self):
        """Get ESP component names."""
        output = list()
        for unit in self.plant.units:
            unit_name = unit.name
            if "esp" in unit_name:
                output.append(unit_name)
        return output

    def get_hexs(self):
        """Get HEX component names."""
        output = list()
        for unit in self.plant.units:
            unit_name = unit.name
            if "heat_exchanger" in unit_name:
                output.append(unit_name)
        return output

    def get_injection_pumps(self):
        """Get injection pump component names."""
        output = list()
        for unit in self.plant.units:
            unit_name = unit.name
            if "injection_pump" in unit_name:
                output.append(unit_name)
        return output

    def get_booster_pumps(self):
        """Get injection pump component names."""
        output = list()
        for unit in self.plant.units:
            unit_name = unit.name
            if "booster_pump" in unit_name:
                output.append(unit_name)
        return output

    def get_aquifers(self):
        """Get aquifer component names."""
        output = list()
        for unit in self.plant.units:
            unit_name = unit.name
            if "aquifer" in unit_name:
                output.append(unit_name)
        return output

    def add_timeseries_plot_to_pdf(self, data, timestamps, xlabel, ylabel, title):
        """Add timeseries plot to PDF."""
        plt.figure(figsize=(10, 5))
        dates = [datetime.fromisoformat(ts.replace("Z", "")) for ts in timestamps]
        plt.plot(timestamps, dates, linestyle="-", color="b", label="Time Series")

        plt.gca().xaxis.set_major_locator(ticker.MaxNLocator(nbins="auto"))
        plt.gca().yaxis.set_major_locator(ticker.MaxNLocator(nbins="auto"))

        plt.xlabel("Timestamp")
        plt.ylabel(ylabel)
        plt.title(title)
        plt.legend()
        plt.grid()

        self.pdf_object.savefig()
        plt.close()

    def add_X_Y_plot_to_pdf(self, x_data, y_data):
        """Add X-Y plot to PDF."""
        plt.figure(figsize=(11.69, 8.27))
        plt.plot(x_data, y_data, marker="o", linestyle="-", color="r", label="X-Y Plot")
        plt.xlabel("X Data")
        plt.ylabel("Y Data")
        plt.title("X-Y Plot")
        plt.legend()
        plt.grid()
        self.pdf_object.savefig()
        plt.close()

    def get_clean_list(self, value_list):
        """Get clean list from value list."""
        clean_list = []
        for value in value_list:
            try:
                # Try converting to float
                converted = float(value)
                clean_list.append(converted)
            except (ValueError, TypeError):
                # Discard if not convertible
                continue
        return clean_list

    def _compose_tagname(self, component_type: str, tagname_value: str) -> str:
        """Return a robust tagname by prefixing a sanitized component_type when needed.

        Some component types use an underscore (e.g. "booster_pump") while stored
        tagnames may omit the underscore ("boosterpump_power..."). This helper
        removes underscores from the component_type and prefixes it to the
        tagname_value when the tagname does not already begin with the
        component_type (with or without underscores).
        """
        if not tagname_value:
            return ""

        ct = (component_type or "")
        sanitized = ct.replace("_", "")

        # Some component types have established short prefixes in tag names
        # (e.g. heat_exchanger -> 'hex'). Provide a small alias map to cover
        # these cases and prefer the alias when prefixing.
        alias_prefix = {
            "heat_exchanger": "hex",
        }

        preferred = alias_prefix.get(ct, sanitized if sanitized else ct)

        # Remove leading underscores from the tag when checking prefixes so that
        # tags like '_power_consumption' are correctly detected as not already
        # prefixed.
        tag_clean = tagname_value.lstrip("_")

        # If tagname already starts with any reasonable prefix, leave it unchanged.
        candidates = {ct, sanitized, preferred}
        for p in candidates:
            if not p:
                continue
            if tag_clean.startswith(p):
                return tagname_value

        # Otherwise prefix the preferred form (alias or sanitized) to the raw
        # tagname (preserving any leading underscore). Example: component_type
        # 'booster_pump' + tagname '_power' -> 'boosterpump_power'.
        return f"{preferred}{tagname_value}"

    def add_stats_plot(self, inj_wells, prod_wells):
        """Add statistics plot to PDF."""
        inj_well_tagnames = [
            "injectionwell_flow.measured",
            "injectionwell_wellhead_pressure.measured",
            "injectionwell_annulus_a_pressure.measured",
        ]
        prod_well_tagnames = [
            "productionwell_annulus_a_pressure.measured",
            "productionwell_annulus_b_pressure.measured",
        ]

        all_wells = inj_wells + prod_wells
        unit_tag_pairs = []

        for well_name in all_wells:
            if "injection" in well_name:
                unit_tag_pairs.append((well_name, inj_well_tagnames))
            elif "production" in well_name:
                unit_tag_pairs.append((well_name, prod_well_tagnames))

        num_units = len(unit_tag_pairs)
        max_tags_per_unit = max(len(tags) for _, tags in unit_tag_pairs)

        # Create subplots with constrained layout
        fig, axes = plt.subplots(
            num_units, max_tags_per_unit, sharex=False, constrained_layout=True
        )

        # Normalize axes to 2D array
        if num_units == 1:
            axes = [axes]
        if max_tags_per_unit == 1:
            axes = [[ax] for ax in axes]

        for row_idx, (well_name, tagnames) in enumerate(unit_tag_pairs):
            self.select_unit(well_name)

            for col_idx in range(max_tags_per_unit):
                ax = axes[row_idx][col_idx]

                if col_idx >= len(tagnames):
                    ax.axis("off")
                    continue

                tagname = tagnames[col_idx]
                value_list, datestamp_list = self.get_data(tagname)
                clean_list = self.get_clean_list(value_list)

                if not clean_list or not datestamp_list:
                    ax.set_title(f"{well_name}\n{tagname}\nNo Data")
                    ax.axis("off")
                    continue

                try:
                    dates = [datetime.fromisoformat(ts.replace("Z", "")) for ts in datestamp_list]
                    max_value = max(clean_list)
                    max_index = value_list.index(max_value)

                    # Plot data
                    ax.plot(dates, value_list, linestyle="-", color="blue", linewidth=1)
                    ax.scatter(dates[max_index], max_value, color="red", zorder=5)

                    # Max value overlay
                    ax.text(
                        0.5,
                        0.5,
                        f"{max_value:.2f}",
                        transform=ax.transAxes,
                        fontsize=26 if num_units > 3 else 32,
                        ha="center",
                        va="center",
                        weight="bold",
                        color="black",
                        zorder=10,
                    )

                    ax.set_title(f"{well_name}\n{tagname}", fontsize=9)
                    ax.grid(True)
                    ax.set_xticks([])
                    ax.set_xticklabels([])
                    ax.tick_params(axis="y", labelsize=8)
                    ax.yaxis.set_major_locator(ticker.MaxNLocator(nbins="auto"))

                except Exception as e:
                    ax.set_title(f"Error: {tagname}")
                    ax.text(0.5, 0.5, str(e), transform=ax.transAxes, ha="center", va="center")
                    ax.axis("off")

        fig.suptitle(
            f"{self.project_name} — Max values during the period "
            f"{self.start_datestamp_title} - {self.end_datestamp_title}",
            fontsize=14,
            fontweight="bold",
        )

        fig.set_size_inches(11.69, 8.27)
        self.pdf_object.savefig(fig, bbox_inches="tight", pad_inches=0.3)
        plt.close(fig)

    def gather_stats(self, well_names, tagnames):
        """Gather statistics for wells."""
        stats_data = []
        for well_name in well_names:
            self.select_unit(well_name)
            for tagname in tagnames:
                value_list, datestamp_list = self.get_data(tagname)
                clean_list = self.get_clean_list(value_list)
                if not clean_list or not datestamp_list:
                    continue  # Skip empty data

                # Compute stats
                max_value = max(clean_list)
                min_value = min(clean_list)
                mean_value = np.mean(clean_list)
                std_value = np.std(clean_list)

                # Timestamp for max value
                try:
                    max_index = value_list.index(max_value)
                    max_timestamp = datestamp_list[max_index]
                    max_datetime = datetime.fromisoformat(max_timestamp.replace("Z", ""))
                    timestamp_str = max_datetime.strftime("%Y-%m-%d %H:%M:%S")
                except Exception:
                    timestamp_str = "N/A"

                stats_data.append(
                    [
                        well_name,
                        tagname,
                        f"{min_value:.2f}",
                        f"{max_value:.2f}",
                        f"{mean_value:.2f}",
                        f"{std_value:.2f}",
                        timestamp_str,
                    ]
                )
        return stats_data

    def add_injection_report(self, inj_wells, tagnames):
        """Add injection report to PDF."""
        num_wells = len(inj_wells)
        if num_wells == 0:
            return

        fig, axes = plt.subplots(num_wells, 1, sharex=True)
        if num_wells == 1:
            axes = [axes]  # Ensure iterable

        fig.suptitle("Injection wells report", fontsize=16, fontweight="bold")
        color_cycle = plt.cm.tab10.colors

        # Separate injectivity tagnames and others, preserving order
        injectivity_tags = [tn for tn in tagnames if "injectivity" in tn.lower()]
        other_tags = [tn for tn in tagnames if "injectivity" not in tn.lower()]

        for idx, (well_name, ax_left) in enumerate(zip(inj_wells, axes)):
            self.select_unit(well_name)

            tag_data = {}
            timestamps = None

            # Collect tag data
            for tagname in tagnames:
                value_list, datestamp_list = self.get_data(tagname)
                if not value_list or not datestamp_list:
                    continue
                try:
                    dates = [datetime.fromisoformat(ts.replace("Z", "")) for ts in datestamp_list]
                    if timestamps is None:
                        timestamps = dates
                    tag_data[tagname] = value_list
                except Exception as e:
                    print(f"Error processing tag '{tagname}' for well '{well_name}': {e}")

            if not tag_data or timestamps is None:
                ax_left.set_title(f"{well_name} (No data)")
                ax_left.axis("off")
                continue

            ax_list = [ax_left]
            color_idx = 0

            # Plot injectivity (left y-axis)
            for tagname in injectivity_tags:
                if tagname in tag_data:
                    label = f"{tagname} [-]"
                    ax_left.plot(
                        timestamps, tag_data[tagname], color=color_cycle[color_idx], label=label
                    )
                    ax_left.set_ylabel(label, color=color_cycle[color_idx])
                    ax_left.tick_params(axis="y", labelcolor=color_cycle[color_idx])
                    ax_left.yaxis.set_major_locator(ticker.MaxNLocator(nbins=5))
                    ax_left.grid(True, linestyle="--", alpha=0.4, axis="y")
                    # Add vertical gridlines
                    ax_left.grid(True, linestyle="--", alpha=0.4, axis="x")
                    color_idx += 1
                    break  # Only one injectivity on left axis
            else:
                # If no injectivity tag found, still add vertical gridlines
                ax_left.grid(True, linestyle="--", alpha=0.4, axis="x")

            # Plot other tags (right y-axes)
            for tagname in other_tags:
                if tagname in tag_data:
                    units = self.get_units(tagname)
                    ax_new = ax_left.twinx()
                    ax_new.spines["right"].set_position(("axes", 1 + 0.1 * (len(ax_list) - 1)))
                    label = f"{tagname} {units}"
                    ax_new.plot(
                        timestamps, tag_data[tagname], color=color_cycle[color_idx], label=label
                    )
                    ax_new.set_ylabel(label, color=color_cycle[color_idx])
                    ax_new.tick_params(axis="y", labelcolor=color_cycle[color_idx])
                    ax_new.yaxis.set_major_locator(ticker.MaxNLocator(nbins=5))
                    ax_new.grid(True, linestyle="--", alpha=0.4, axis="y")
                    # Add vertical gridlines for right axes as well (optional, usually not needed)
                    ax_new.grid(True, linestyle="--", alpha=0.4, axis="x")
                    ax_list.append(ax_new)
                    color_idx += 1

            ax_left.set_title(f"{well_name}", fontsize=10)
            ax_left.xaxis.set_major_locator(mdates.AutoDateLocator())
            ax_left.xaxis.set_major_formatter(mdates.DateFormatter("%m/%d %H:%M"))
            ax_left.tick_params(axis="x", labelrotation=45, labelsize=8)

        # Default to A4 landscape in inches
        width, height = 11.69, 8.27
        fig.set_size_inches(width, height)
        self.pdf_object.savefig(fig, bbox_inches="tight")
        plt.close(fig)

    def add_production_report(self, prod_wells, tagnames):
        """Add production report to PDF."""
        num_wells = len(prod_wells)
        if num_wells == 0:
            return

        fig, axes = plt.subplots(num_wells, 1, sharex=True)
        if num_wells == 1:
            axes = [axes]  # Ensure iterable

        fig.suptitle("Production wells report", fontsize=16, fontweight="bold")
        color_cycle = plt.cm.tab10.colors

        # Separate annulus pressure tags and others, preserving order
        annulus_p_tags = [tn for tn in tagnames if "annulus" in tn.lower()]
        other_tags = [tn for tn in tagnames if "annulus" not in tn.lower()]

        for idx, (well_name, ax_left) in enumerate(zip(prod_wells, axes)):
            self.select_unit(well_name)

            tag_data = {}
            timestamps = None

            # Collect tag data
            for tagname in tagnames:
                value_list, datestamp_list = self.get_data(tagname)
                if not value_list or not datestamp_list:
                    continue
                try:
                    dates = [datetime.fromisoformat(ts.replace("Z", "")) for ts in datestamp_list]
                    if timestamps is None:
                        timestamps = dates
                    tag_data[tagname] = value_list
                except Exception as e:
                    print(f"Error processing tag '{tagname}' for well '{well_name}': {e}")

            if not tag_data or timestamps is None:
                ax_left.set_title(f"{well_name} (No data)")
                ax_left.axis("off")
                continue

            ax_list = [ax_left]
            color_idx = 0

            # Plot annulus pressures (left y-axis)
            for tagname in annulus_p_tags:
                if tagname in tag_data:
                    # Determine units for annulus tags, default to [bar]
                    units = self.get_units(tagname)
                    label = f"{tagname} {units}"
                    ax_left.plot(
                        timestamps, tag_data[tagname], color=color_cycle[color_idx], label=label
                    )
                    ax_left.set_ylabel(label, color=color_cycle[color_idx])
                    ax_left.tick_params(axis="y", labelcolor=color_cycle[color_idx])
                    ax_left.yaxis.set_major_locator(ticker.MaxNLocator(nbins=5))
                    ax_left.grid(True, linestyle="--", alpha=0.4, axis="y")
                    ax_left.grid(True, linestyle="--", alpha=0.4, axis="x")  # vertical gridlines
                    color_idx += 1
            else:
                # If no annulus pressure tag found, still add vertical gridlines on left axis
                ax_left.grid(True, linestyle="--", alpha=0.4, axis="x")

            # Plot other tags (right y-axes)
            for tagname in other_tags:
                if tagname in tag_data:
                    # Units for other tags, extend if needed
                    units = self.get_units(tagname)
                    ax_new = ax_left.twinx()
                    ax_new.spines["right"].set_position(("axes", 1 + 0.1 * (len(ax_list) - 1)))
                    label = f"{tagname} {units}"
                    ax_new.plot(
                        timestamps, tag_data[tagname], color=color_cycle[color_idx], label=label
                    )
                    ax_new.set_ylabel(label, color=color_cycle[color_idx])
                    ax_new.tick_params(axis="y", labelcolor=color_cycle[color_idx])
                    ax_new.yaxis.set_major_locator(ticker.MaxNLocator(nbins=5))
                    ax_new.grid(True, linestyle="--", alpha=0.4, axis="y")
                    ax_new.grid(True, linestyle="--", alpha=0.4, axis="x")
                    ax_list.append(ax_new)
                    color_idx += 1

            ax_left.set_title(f"{well_name}", fontsize=10)
            ax_left.xaxis.set_major_locator(mdates.AutoDateLocator())
            ax_left.xaxis.set_major_formatter(mdates.DateFormatter("%m/%d %H:%M"))
            ax_left.tick_params(axis="x", labelrotation=45, labelsize=8)

        width, height = 11.69, 8.27
        fig.set_size_inches(width, height)
        self.pdf_object.savefig(fig, bbox_inches="tight")
        plt.close(fig)

    def add_esp_report(self, esps, options):
        """Add ESP report to PDF."""
        rcParams["figure.figsize"] = [11.69, 8.27]  # A4 landscape
        color_cycle = plt.cm.tab10.colors
        plots_per_page = 6  # 2 columns x 3 rows
        ncols = 2
        nrows = 3

        for esp in esps:
            self.select_unit(esp)

            selected_plots = [
                (key, opt["tagname"], opt.get("min"), opt.get("max"))
                for key, opt in options.items()
                if opt.get("checked") and "tagname" in opt
            ]

            total_plots = len(selected_plots)
            total_pages = ceil(total_plots / plots_per_page)

            for page_index in range(total_pages):
                fig, axes = plt.subplots(nrows=nrows, ncols=ncols, sharex=True)
                axes = axes.flatten()

                fig.suptitle(
                    f"ESP Report - {esp} - Page {page_index + 1}", fontsize=16, fontweight="bold"
                )
                fig.subplots_adjust(
                    left=0.06, right=0.95, top=0.88, bottom=0.10, wspace=0.2, hspace=0.5
                )

                for i in range(plots_per_page):
                    subplot_index = page_index * plots_per_page + i
                    if subplot_index >= total_plots:
                        axes[i].axis("off")
                        continue

                    key, tagname, min_val, max_val = selected_plots[subplot_index]
                    ax = axes[i]

                    value_list, datestamp_list = self.get_data(tagname)
                    if not value_list or not datestamp_list:
                        ax.set_title(f"{tagname} (No data)")
                        ax.axis("off")
                        continue

                    try:
                        timestamps = [
                            datetime.fromisoformat(ts.replace("Z", "")) for ts in datestamp_list
                        ]
                    except Exception as e:
                        print(f"Error processing tag '{tagname}' for ESP '{esp}': {e}")
                        ax.set_title(f"{tagname} (Error)")
                        ax.axis("off")
                        continue

                    units = self.get_units(tagname)
                    color = color_cycle[i % len(color_cycle)]

                    ax.plot(timestamps, value_list, color=color)
                    ax.tick_params(axis="y", labelcolor=color)
                    ax.yaxis.set_major_locator(ticker.MaxNLocator(nbins=5))
                    ax.grid(True, linestyle="--", alpha=0.4, axis="both")

                    # Add min/max lines and fill regions
                    try:
                        if min_val is not None:
                            min_float = float(min_val)
                            ax.axhline(y=min_float, color="green", linestyle="dotted", linewidth=1)
                            ax.fill_between(
                                timestamps, ax.get_ylim()[0], min_float, color="#d6f5d6", alpha=0.4
                            )
                    except ValueError:
                        pass  # Ignore invalid min_val

                    try:
                        if max_val is not None:
                            max_float = float(max_val)
                            ax.axhline(y=max_float, color="green", linestyle="dotted", linewidth=1)
                            ax.fill_between(
                                timestamps, max_float, ax.get_ylim()[1], color="#d6f5d6", alpha=0.4
                            )
                    except ValueError:
                        pass  # Ignore invalid max_val

                    ax.set_title(f"{tagname} {units}", fontsize=10)
                    ax.set_xlabel("Date")
                    ax.xaxis.set_major_locator(mdates.AutoDateLocator())
                    ax.xaxis.set_major_formatter(mdates.DateFormatter("%m/%d"))
                    ax.tick_params(axis="x", labelrotation=45, labelsize=8)

                for j in range(len(selected_plots) % plots_per_page, plots_per_page):
                    if (page_index * plots_per_page + j) >= total_plots:
                        axes[j].axis("off")

                for ax in axes:
                    ax.tick_params(axis="x", labelbottom=True, pad=2)

                fig.set_size_inches(11.69, 8.27)  # A4 landscape
                self.pdf_object.savefig(fig, bbox_inches="tight")
                plt.close(fig)

    def add_stats_table(self, inj_wells, prod_wells):
        """Add statistics table to PDF."""
        all_stats_data = []

        # Tagnames for values to be added in the stats table
        inj_well_tagnames = [
            "injectionwell_flow.measured",
            "injectionwell_wellhead_pressure.measured",
            "injectionwell_annulus_pressure.measured",
        ]
        prod_well_tagnames = ["productionwell_annulus_a_pressure.measured"]

        # First add injection well stats, then production well stats
        all_stats_data += self.gather_stats(inj_wells, inj_well_tagnames)
        all_stats_data += self.gather_stats(prod_wells, prod_well_tagnames)

        if not all_stats_data:
            return  # Skip if no valid stats found

        # Create DataFrame
        df = pd.DataFrame(
            all_stats_data,
            columns=[
                "Unit Name",
                "Tag Name",
                "Min Value",
                "Max Value",
                "Mean Value",
                "Std Dev",
                "Timestamp of Max",
            ],
        )

        # Create figure and table
        fig, ax = plt.subplots()
        ax.axis("tight")
        ax.axis("off")

        ax.set_title(
            f"Summary table for the period {self.start_datestamp_title} - "
            f"{self.end_datestamp_title}",
            fontsize=14,
            fontweight="bold",
            pad=20,
        )

        table = ax.table(
            cellText=df.values,
            colLabels=df.columns,
            cellLoc="center",
            loc="center",
            colColours=["lightgray"] * df.shape[1],
        )

        table.auto_set_font_size(False)
        table.set_fontsize(9)
        table.auto_set_column_width(col=list(range(len(df.columns))))

        # Save figure to PDF
        width, height = 11.69, 8.27
        fig.set_size_inches(width, height)
        self.pdf_object.savefig(fig, bbox_inches="tight", pad_inches=0.5)
        plt.close(fig)

    def add_cross_plot(self, units, tagnames, plot_type):
        """Add cross plot to PDF."""
        num_units = len(units)
        if num_units == 0 or len(tagnames) < 3:
            print("Insufficient input: need at least one unit and three tagnames.")
            return

        # Layout: Try to form a near-square grid
        ncols = ceil(sqrt(num_units))
        nrows = ceil(num_units / ncols)

        fig, axes = plt.subplots(nrows, ncols, sharex=False, sharey=False, constrained_layout=True)

        # Normalize axes to 2D list
        axes = np.array(axes).reshape(nrows, ncols)

        for idx, unit_name in enumerate(units):
            row, col = divmod(idx, ncols)
            ax = axes[row][col]

            self.select_unit(unit_name)
            tagname_y = tagnames[0]
            tagname_x = tagnames[1]
            tagname_z = tagnames[2]

            y_data, _ = self.get_data(tagname_y)
            x_data, _ = self.get_data(tagname_x)

            if not y_data or not x_data:
                print(f"[{unit_name}] Missing data for x or y tag.")
                ax.set_title(f"{unit_name} (No data)")
                ax.axis("off")
                continue

            if tagname_z.lower() == "datestamp":
                _, datestamps_z = self.get_data(tagname_y)  # shared timestamps
                try:
                    z_datetimes = [
                        datetime.fromisoformat(ts.replace("Z", "")) for ts in datestamps_z
                    ]
                    z_data = [dt.timestamp() for dt in z_datetimes]
                    z_units = ""
                except Exception as e:
                    print(f"[{unit_name}] Failed to parse datestamp for z: {e}")
                    ax.set_title(f"{unit_name} (Invalid datestamp)")
                    ax.axis("off")
                    continue
            else:
                z_data, _ = self.get_data(tagname_z)
                if not z_data:
                    print(f"[{unit_name}] Missing data for z tag.")
                    ax.set_title(f"{unit_name} (No z data)")
                    ax.axis("off")
                    continue
                z_units = self.get_units(tagname_z)
                z_datetimes = None

            x_units = self.get_units(tagname_x)
            y_units = self.get_units(tagname_y)

            # Scatter plot
            scatter = ax.scatter(x_data, y_data, c=z_data, cmap="viridis", edgecolors="none")
            ax.set_title(f"{unit_name}", fontsize=10)
            ax.set_xlabel(f"{tagname_x} {x_units}", fontsize=8)
            ax.set_ylabel(f"{tagname_y} {y_units}", fontsize=8)
            ax.tick_params(labelsize=8)

            # Colorbar
            cbar = fig.colorbar(scatter, ax=ax)
            if tagname_z.lower() == "datestamp" and z_datetimes:
                formatter = ticker.FuncFormatter(
                    lambda val, pos: datetime.fromtimestamp(val).strftime("%m/%d %H:%M")
                )
                cbar.ax.yaxis.set_major_formatter(formatter)
                cbar.set_label(f"{tagname_z}", fontsize=8)
            else:
                cbar.set_label(f"{tagname_z} {z_units}", fontsize=8)

        # Turn off unused subplots
        total_axes = nrows * ncols
        if total_axes > num_units:
            for empty_idx in range(num_units, total_axes):
                row, col = divmod(empty_idx, ncols)
                axes[row][col].axis("off")

        # Title
        fig.suptitle(f"{plot_type} Cross Plots", fontsize=14, fontweight="bold")

        # A4 landscape size
        fig.set_size_inches(11.69, 8.27)
        self.pdf_object.savefig(fig, bbox_inches="tight", pad_inches=0.3)
        plt.close(fig)

    def compute_skin_lines(self, inputs, flow_array, skin_array):
        """Compute skin lines for plot."""
        well_name = inputs["well_name"]
        print(f"Computing skin lines for well: {well_name}")

        app_IWM = InjectionWellMonitoring()
        app_IWM.load_plant(self.project_path, self.project_name)
        app_IWM.select_unit(well_name)

        print(app_IWM.unit.name)
        print(app_IWM.unit.to_units[0].name)

        well_param = app_IWM.unit.parameters["property"]
        reservoir_param = app_IWM.unit.to_units[0].parameters["property"]

        # Get inputs for skin lines
        boundary = {
            "min_flow_plot": inputs["min_flow_plot"],
            "max_flow_plot": inputs["max_flow_plot"],
            "no_interval_flow_plot": inputs["no_interval_flow_plot"],
            "min_skin_plot": inputs["min_skin_plot"],
            "max_skin_plot": inputs["max_skin_plot"],
            "no_interval_skin_plot": inputs["no_interval_skin_plot"],
            "max_pressure": None,
            "max_flow_rate": None,
            "wellbore_radius": well_param["wellbore_radius"][0],
            "start_time": inputs["starttime"],
            "end_time": inputs["endtime"],
        }
        app_IWM.set_input(boundary)

        parameters = {
            "reservoir_pressure": reservoir_param["reservoir_pressure"][0],
            "reservoir_radius": reservoir_param["reservoir_radius"][0],
            "reservoir_permeability": reservoir_param["reservoir_permeability"][0],
            "reservoir_thickness": reservoir_param["reservoir_thickness"][0],
            "reservoir_top": reservoir_param["reservoir_top"][0],
            "liquid_density": reservoir_param["liquid_density"][0],
            "liquid_viscosity": reservoir_param["liquid_viscosity"][0],
        }
        app_IWM.init_parameters(parameters)
        app_IWM.get_data()
        app_IWM.calculate_skin_lines()

        inputs = app_IWM.get_input()
        outputs = app_IWM.get_output()

        results = {
            "injection_pressure": outputs["injection_pressure"],
            "max_cal_P_inj": outputs["max_cal_P_inj"],
        }
        return results

    def convert_numeric_values(self, inputs):
        """Convert numeric values in inputs."""
        for key, val in inputs.items():
            # Try to convert to float first
            try:
                num = float(val)
                # If no error, check if it can be an int
                if num.is_integer():
                    inputs[key] = int(num)
                else:
                    inputs[key] = num
            except (ValueError, TypeError):
                # Not a number, leave as is
                pass
        return inputs

    def add_cross_plot_with_skin_lines(self, units, tagnames, inputs):
        """Add cross plot with skin lines to PDF."""
        # Make sure numeric values are not in string format
        inputs = self.convert_numeric_values(inputs)

        num_units = len(units)
        if num_units == 0 or len(tagnames) < 3:
            print("Insufficient input: need at least one unit and three tagnames.")
            return

        ncols = ceil(sqrt(num_units))
        nrows = ceil(num_units / ncols)

        fig, axes = plt.subplots(nrows, ncols, sharex=False, sharey=False, constrained_layout=True)
        axes = np.array(axes).reshape(nrows, ncols)

        cmap = plt.cm.plasma

        colors = [
            cmap(i / (inputs["no_interval_skin_plot"] - 1))
            for i in range(inputs["no_interval_skin_plot"])
        ]

        for idx, unit_name in enumerate(units):
            row, col = divmod(idx, ncols)
            ax = axes[row][col]

            self.select_unit(unit_name)

            tagname_y, tagname_x, tagname_z = tagnames[:3]

            y_data, _ = self.get_data(tagname_y)  # Pressure in bar
            x_data, _ = self.get_data(tagname_x)  # Flow in m³/h

            if not y_data or not x_data:
                print(f"[{unit_name}] Missing data for x or y tag.")
                ax.set_title(f"{unit_name} (No data)")
                ax.axis("off")
                continue

            skin_array = np.linspace(
                inputs["min_skin_plot"], inputs["max_skin_plot"], inputs["no_interval_skin_plot"]
            )
            flow_array = np.linspace(
                inputs["min_flow_plot"] / 3600,
                inputs["max_flow_plot"] / 3600,
                inputs["no_interval_flow_plot"],
            )

            # Compute skin lines matrix
            inputs["well_name"] = unit_name
            skin_results = self.compute_skin_lines(inputs, flow_array, skin_array)
            pressure_matrix = skin_results["injection_pressure"]

            # Z data for coloring scatter
            if tagname_z.lower() == "datestamp":
                _, datestamps_z = self.get_data(tagname_y)
                try:
                    z_datetimes = [
                        datetime.fromisoformat(ts.replace("Z", "")) for ts in datestamps_z
                    ]
                    z_data = [dt.timestamp() for dt in z_datetimes]
                    z_units = ""
                except Exception as e:
                    print(f"[{unit_name}] Failed to parse datestamp for z: {e}")
                    ax.set_title(f"{unit_name} (Invalid datestamp)")
                    ax.axis("off")
                    continue
            else:
                z_data, _ = self.get_data(tagname_z)
                if not z_data:
                    print(f"[{unit_name}] Missing data for z tag.")
                    ax.set_title(f"{unit_name} (No z data)")
                    ax.axis("off")
                    continue
                z_units = self.get_units(tagname_z)

            x_units = self.get_units(tagname_x)
            y_units = self.get_units(tagname_y)

            # Scatter plot
            scatter = ax.scatter(x_data, y_data, c=z_data, cmap="viridis", edgecolors="none")
            ax.set_title(f"{unit_name}", fontsize=10)
            ax.set_xlabel(f"{tagname_x} {x_units}", fontsize=8)
            ax.set_ylabel(f"{tagname_y} {y_units}", fontsize=8)
            ax.tick_params(labelsize=8)

            # Plot skin lines (flow in m³/h, pressure in bar)
            flow_hr = flow_array * 3600
            for i, skin in enumerate(skin_array):
                ax.plot(
                    flow_hr,
                    pressure_matrix[i],
                    color=colors[i],
                    linestyle="dotted",
                    linewidth=1,
                    label=f"Skin = {skin}",
                )

            ax.legend(fontsize=8, title="Skin Values")

            # Colorbar
            cbar = fig.colorbar(scatter, ax=ax)
            if tagname_z.lower() == "datestamp":
                formatter = ticker.FuncFormatter(
                    lambda val, pos: datetime.fromtimestamp(val).strftime("%m/%d %H:%M")
                )
                cbar.ax.yaxis.set_major_formatter(formatter)
            cbar.set_label(f"{tagname_z} {z_units}", fontsize=8)

        # Hide unused axes
        total_axes = nrows * ncols
        if total_axes > num_units:
            for empty_idx in range(num_units, total_axes):
                row, col = divmod(empty_idx, ncols)
                axes[row][col].axis("off")

        fig.suptitle(
            f"{inputs['plot_type']} Cross Plots with Skin Lines", fontsize=14, fontweight="bold"
        )
        fig.set_size_inches(11.69, 8.27)  # A4 landscape
        self.pdf_object.savefig(fig, bbox_inches="tight", pad_inches=0.3)
        plt.close(fig)

    def calculate_total_volume(self, flow_rates, timestamps):
        """Calculate total volume from timestamps and flow rates."""
        if len(timestamps) != len(flow_rates):
            raise ValueError("timestamps and flow_rates must have the same length")

        if len(timestamps) < 2:
            return 0.0, []  # not enough data

        total_volume = 0.0
        interval_volumes = []

        # Convert timestamps into datetime objects
        times = [datetime.strptime(ts.replace("Z", ""), "%Y-%m-%dT%H:%M:%S") for ts in timestamps]

        # Initialize previous valid value
        prev_flow = None

        for i in range(len(times) - 1):
            f0 = flow_rates[i]
            f1 = flow_rates[i + 1]

            # Replace None with previous valid value
            if f0 is None:
                if prev_flow is None:
                    f0 = 0.0  # fallback if first value is None
                else:
                    f0 = prev_flow

            if f1 is None:
                f1 = f0  # use f0 if f1 is None

            # Average flow during this interval
            avg_flow = (f0 + f1) / 2.0

            # Time difference in hours
            delta_hours = (times[i + 1] - times[i]).total_seconds() / 3600.0

            # Interval volume
            interval_volume = avg_flow * delta_hours

            interval_volumes.append(interval_volume)
            total_volume += interval_volume

            # Update prev_flow
            prev_flow = f1

        return total_volume, interval_volumes

    def weighted_average_value_with_volume(self, values, timestamps, volume_intervals):
        """Calculate the volume-weighted average of a value over time."""
        if len(values) != len(timestamps):
            raise ValueError("values and timestamps must have the same length")

        if len(volume_intervals) != len(timestamps) - 1:
            raise ValueError("volume_intervals length must be len(timestamps)-1")

        if len(volume_intervals) == 0:
            return 0.0  # nothing to average

        # -------------------------
        # Interpolate None values
        # -------------------------
        values_interp = values.copy()
        n = len(values_interp)

        # Find indices of valid values
        valid_indices = [i for i, v in enumerate(values_interp) if v is not None]

        if not valid_indices:
            return 0.0  # all values are None

        for i in range(n):
            if values_interp[i] is None:
                # find previous and next valid indices
                prev_idx = max([vi for vi in valid_indices if vi < i], default=None)
                next_idx = min([vi for vi in valid_indices if vi > i], default=None)

                if prev_idx is None:
                    # Use next valid value
                    values_interp[i] = values_interp[next_idx]
                elif next_idx is None:
                    # Use previous valid value
                    values_interp[i] = values_interp[prev_idx]
                else:
                    # Linear interpolation
                    prev_val = values_interp[prev_idx]
                    next_val = values_interp[next_idx]
                    values_interp[i] = prev_val + (next_val - prev_val) * (i - prev_idx) / (
                        next_idx - prev_idx
                    )

        # -------------------------
        # Compute weighted average
        # -------------------------
        weighted_sum = 0.0
        total_volume = 0.0

        for i in range(len(volume_intervals)):
            # average value in this interval
            avg_value = (values_interp[i] + values_interp[i + 1]) / 2.0

            # multiply by interval volume
            weighted_sum += avg_value * volume_intervals[i]

            # sum total volume
            total_volume += volume_intervals[i]

        if total_volume == 0:
            return 0.0

        return weighted_sum / total_volume

    def align_series_on_common_timestamps(self, values_a, timestamps_a, values_b, timestamps_b):
        """Align two series on shared timestamps, preserving series A ordering."""
        if values_a is None or timestamps_a is None or values_b is None or timestamps_b is None:
            return [], [], []

        series_b = {ts: val for ts, val in zip(timestamps_b, values_b)}
        common_timestamps = [ts for ts in timestamps_a if ts in series_b]

        aligned_a = [values_a[i] for i, ts in enumerate(timestamps_a) if ts in series_b]
        aligned_b = [series_b[ts] for ts in common_timestamps]

        return aligned_a, aligned_b, common_timestamps

    def format_value(self, val):
        """Change the format of a value."""
        if isinstance(val, (float, int)):
            return f"{val:.2f}"
        else:
            return str(val)

    def find_adjacent_component_name(self, main_unit, adj_unit):
        """Return the adjacent component name matching the requested unit type."""
        self.select_unit(main_unit)
        for unit in self.unit.to_units:
            if adj_unit in unit.name:
                self.select_unit(unit.name)
                return unit.name

    @staticmethod
    def _is_nlog_row_enabled(row):
        """Return True unless an NLOG row explicitly disables the parameter."""
        value = row.get("enabled", True)
        if isinstance(value, str):
            return value.strip().lower() not in {"false", "0", "no", "off"}
        return value is not False

    def _enabled_nlog_rows(self, rows):
        return [row for row in (rows or []) if self._is_nlog_row_enabled(row)]

    @staticmethod
    def _ensure_nlog_enabled_default(rows):
        for row in rows or []:
            row.setdefault("enabled", True)
        return rows

    def calculate_nlog_1A(self, rows, prod_wells, esps):
        """Calculate NLOG Section 1A values for each production well."""
        all_rows = rows or []
        enabled_parameters = {
            row.get("parameter") for row in all_rows if self._is_nlog_row_enabled(row)
        }
        rows = self._enabled_nlog_rows(all_rows)

        # Output structure
        result_rows = []

        # Ordered parameter list
        param_order = [
            "prod_vol_water",
            "prod_temp_avg_weighted",
            "prod_pres_avg",
            "prod_pres_min",
            "prod_wh_pres",
            "prod_oil_vol",
            "prod_gas_vol",
            "prod_condens_vol",
            "prod_inhibit_vol",
        ]

        # Convert rows list into lookup table:
        # lookup[(component_name, parameter)] -> tagname
        tag_lookup = {(r["component_name"], r["parameter"]): r["tagname"] for r in rows}

        # Get unique production wells from rows (respects user-selected units from frontend)
        selected_prod_wells = list(
            set(r["component_name"] for r in rows if r.get("component_type") == "production_well")
        )
        if not selected_prod_wells:
            selected_prod_wells = prod_wells  # Fallback to defaults if none in rows

        # Loop through selected production wells
        for well_name in selected_prod_wells:
            self.select_unit(well_name)

            # Find ESP and aquifer for this production well
            # First try to find from rows (user-selected), then fallback to system discovery
            esp_name = next(
                (r["component_name"] for r in rows if r.get("component_type") == "esp"),
                self.find_adjacent_component_name(well_name, "esp"),
            )
            aquifer_name = next(
                (r["component_name"] for r in rows if r.get("component_type") == "aquifer"),
                self.find_adjacent_component_name(well_name, "aquifer"),
            )
            # Prepare a row output
            # 1
            tagname_prod_vol = tag_lookup.get((well_name, "prod_vol_water"))
            if tagname_prod_vol:
                flow_values, flow_datestamps = self.get_data(tagname_prod_vol)
                prod_vol_water, _ = self.calculate_total_volume(flow_values, flow_datestamps)
            else:
                flow_values, flow_datestamps = None, None
                prod_vol_water = None

            # 2
            # Calculate volume_intervals from flow for use in all flow-dependent averages
            _, flow_volume_intervals = (
                self.calculate_total_volume(flow_values, flow_datestamps)
                if flow_values
                else (None, [])
            )

            tagname_prod_temp = tag_lookup.get((well_name, "prod_temp_avg_weighted"))
            values, datestamps = (
                self.get_data(tagname_prod_temp) if tagname_prod_temp else (None, None)
            )
            if (
                flow_values is not None
                and flow_datestamps is not None
                and values is not None
                and datestamps is not None
            ):
                aligned_flow, aligned_values, aligned_datestamps = (
                    self.align_series_on_common_timestamps(
                        flow_values,
                        flow_datestamps,
                        values,
                        datestamps,
                    )
                )
                if len(aligned_datestamps) > 1:
                    _, volume_intervals = self.calculate_total_volume(
                        aligned_flow, aligned_datestamps
                    )
                    prod_temp_avg_weighted = self.weighted_average_value_with_volume(
                        aligned_values, aligned_datestamps, volume_intervals
                    )
                else:
                    prod_temp_avg_weighted = None
            else:
                prod_temp_avg_weighted = None

            # 5
            # Wellhead pressure - use flow-weighted averages instead of uniform weights
            tagname_prod_wh = tag_lookup.get((well_name, "prod_wh_pres"))
            values, datestamps = self.get_data(tagname_prod_wh) if tagname_prod_wh else (None, None)
            if (
                values is not None
                and datestamps is not None
                and len(values) > 1
                and len(values) == len(datestamps)
                and flow_volume_intervals
            ):
                # Align pressure with flow timestamps
                aligned_flow_wh, aligned_pres_wh, aligned_ts_wh = (
                    self.align_series_on_common_timestamps(
                        flow_values, flow_datestamps, values, datestamps
                    )
                )
                if len(aligned_ts_wh) > 1:
                    _, volume_intervals_wh = self.calculate_total_volume(
                        aligned_flow_wh, aligned_ts_wh
                    )
                    prod_wh_pres = self.weighted_average_value_with_volume(
                        aligned_pres_wh, aligned_ts_wh, volume_intervals_wh
                    )
                else:
                    prod_wh_pres = None
            else:
                prod_wh_pres = None

            # 3
            # Production pressure from ESP - use flow-weighted averages instead of uniform weights
            self.select_unit(esp_name)
            tagname_esp_pres_avg = tag_lookup.get((esp_name, "prod_pres_avg"))
            values, datestamps = (
                self.get_data(tagname_esp_pres_avg) if tagname_esp_pres_avg else (None, None)
            )
            if values is not None and len(values) > 1 and flow_volume_intervals:
                # Align ESP pressure with flow timestamps
                aligned_flow_esp, aligned_pres_esp, aligned_ts_esp = (
                    self.align_series_on_common_timestamps(
                        flow_values, flow_datestamps, values, datestamps
                    )
                )
                if len(aligned_ts_esp) > 1:
                    _, volume_intervals_esp = self.calculate_total_volume(
                        aligned_flow_esp, aligned_ts_esp
                    )
                    prod_pres_avg = self.weighted_average_value_with_volume(
                        aligned_pres_esp, aligned_ts_esp, volume_intervals_esp
                    )
                else:
                    prod_pres_avg = None
            else:
                prod_pres_avg = None

            # 4
            tagname_esp_pres_min = tag_lookup.get((esp_name, "prod_pres_min"))
            values, _ = (
                self.get_data(tagname_esp_pres_min) if tagname_esp_pres_min else (None, None)
            )
            if values is not None and len(values) > 1:
                prod_pres_min = min(values)
            else:
                prod_pres_min = None

            if aquifer_name is not None:
                # 6: prod_oil_vol
                tagname_oil = tag_lookup.get((aquifer_name, "prod_oil_vol"))
                if tagname_oil:
                    values, datestamps = self.get_data(tagname_oil)
                    prod_oil_vol, _ = self.calculate_total_volume(values, datestamps)
                else:
                    prod_oil_vol = None

                # 7: prod_gas_vol
                tagname_gas = tag_lookup.get((aquifer_name, "prod_gas_vol"))
                if tagname_gas:
                    values, datestamps = self.get_data(tagname_gas)
                    prod_gas_vol, _ = self.calculate_total_volume(values, datestamps)
                else:
                    prod_gas_vol = None

                # 8: prod_condens_vol
                tagname_cond = tag_lookup.get((aquifer_name, "prod_condens_vol"))
                if tagname_cond:
                    values, datestamps = self.get_data(tagname_cond)
                    prod_condens_vol, _ = self.calculate_total_volume(values, datestamps)
                else:
                    prod_condens_vol = None

                # 9: prod_inhibit_vol
                tagname_inhibit = tag_lookup.get((aquifer_name, "prod_inhibit_vol"))
                if tagname_inhibit:
                    values, datestamps = self.get_data(tagname_inhibit)
                    prod_inhibit_vol, _ = self.calculate_total_volume(values, datestamps)
                else:
                    prod_inhibit_vol = None
            else:
                prod_oil_vol = None
                prod_gas_vol = None
                prod_condens_vol = None
                prod_inhibit_vol = None

            row_data = {
                "well_name": well_name,
                "prod_vol_water": prod_vol_water,
                "prod_temp_avg_weighted": prod_temp_avg_weighted,
                "prod_pres_avg": prod_pres_avg,
                "prod_pres_min": prod_pres_min,
                "prod_wh_pres": prod_wh_pres,
                "prod_oil_vol": prod_oil_vol,
                "prod_gas_vol": prod_gas_vol,
                "prod_condens_vol": prod_condens_vol,
                "prod_inhibit_vol": prod_inhibit_vol,
            }

            # Add row to results
            result_rows.append(row_data)

        # Convert to dataframe in correct column order
        df = pd.DataFrame(result_rows, columns=["well_name"] + param_order)
        for parameter in param_order:
            if parameter not in enabled_parameters:
                df[parameter] = None

        return df

    import pandas as pd

    def calculate_nlog_1B(self, rows, inj_wells, hexs, injection_pumps):
        """
        Calculate NLOG Section 1B values for each injection well.

        Output columns:
            - well_name
            - inj_vol_water
            - inj_temp_avg_weighted
            - inj_pres_avg
            - inj_pres_max
            - inj_inhibit_vol
        """
        all_rows = rows or []
        enabled_parameters = {
            row.get("parameter") for row in all_rows if self._is_nlog_row_enabled(row)
        }
        rows = self._enabled_nlog_rows(all_rows)

        # Output structure
        result_rows = []

        # Ordered parameter list
        param_order = [
            "inj_vol_water",
            "inj_temp_avg_weighted",
            "inj_pres_avg",
            "inj_pres_max",
            "inj_inhibit_vol",
        ]

        # Build lookup: (component_name, parameter) -> tagname
        tag_lookup = {(r["component_name"], r["parameter"]): r["tagname"] for r in rows}

        # Get unique injection wells from rows (respects user-selected units from frontend)
        selected_inj_wells = list(
            set(r["component_name"] for r in rows if r.get("component_type") == "injection_well")
        )
        if not selected_inj_wells:
            selected_inj_wells = inj_wells  # Fallback to defaults if none in rows

        # Loop through selected injection wells
        for ii, well_name in enumerate(selected_inj_wells):

            # Find HEX and injection pump for this injection well
            hex_name = next(
                (r["component_name"] for r in rows if r.get("component_type") == "heat_exchanger"),
                None,
            )
            if hex_name is None:
                hex_name = self.find_adjacent_component_name(well_name, "heat_exchanger")
            if hex_name is None and len(hexs) > 0:
                hex_name = hexs[ii]

            inj_pump_name = next(
                (r["component_name"] for r in rows if r.get("component_type") == "injection_pump"),
                None,
            )
            if inj_pump_name is None:
                inj_pump_name = self.find_adjacent_component_name(well_name, "injection_pump")
            if inj_pump_name is None and len(injection_pumps) > 0:
                inj_pump_name = injection_pumps[ii]
            # 1) inj_vol_water (total injected water volume)
            self.select_unit(well_name)
            tagname_inj_flow = tag_lookup.get((well_name, "inj_vol_water"))
            flow_values, flow_datestamps = (
                self.get_data(tagname_inj_flow) if tagname_inj_flow else (None, None)
            )
            if flow_values is not None and flow_datestamps is not None:
                inj_vol_water, _ = self.calculate_total_volume(flow_values, flow_datestamps)
            else:
                inj_vol_water = None
                flow_values, flow_datestamps = None, None

            # 2) inj_temp_avg_weighted (weighted average of HEX primary outlet temperature)
            inj_temp_avg_weighted = None
            if hex_name is not None:
                self.select_unit(hex_name)
                tagname_inj_temp = tag_lookup.get((hex_name, "inj_temp_avg_weighted"))
                values, datestamps = (
                    self.get_data(tagname_inj_temp) if tagname_inj_temp else (None, None)
                )
                if (
                    flow_values is not None
                    and flow_datestamps is not None
                    and values is not None
                    and datestamps is not None
                ):
                    aligned_flow, aligned_values, aligned_datestamps = (
                        self.align_series_on_common_timestamps(
                            flow_values,
                            flow_datestamps,
                            values,
                            datestamps,
                        )
                    )
                    if len(aligned_datestamps) > 1:
                        _, volume_intervals = self.calculate_total_volume(
                            aligned_flow, aligned_datestamps
                        )
                        inj_temp_avg_weighted = self.weighted_average_value_with_volume(
                            aligned_values, aligned_datestamps, volume_intervals
                        )

            # 3) inj_pres_avg: average outlet pressure from injection pump.
            # Use flow-weighted averages.
            inj_pres_avg = None
            if (
                inj_pump_name is not None
                and flow_values is not None
                and flow_datestamps is not None
            ):
                self.select_unit(inj_pump_name)
                tagname_inj_pres_avg = tag_lookup.get((inj_pump_name, "inj_pres_avg"))
                values, datestamps = (
                    self.get_data(tagname_inj_pres_avg) if tagname_inj_pres_avg else (None, None)
                )
                if (
                    values is not None
                    and datestamps is not None
                    and len(values) > 0
                    and len(datestamps) > 0
                ):
                    # Align pump pressure with flow timestamps for flow-weighted averaging
                    aligned_flow_pres, aligned_pres, aligned_ts_pres = (
                        self.align_series_on_common_timestamps(
                            flow_values, flow_datestamps, values, datestamps
                        )
                    )
                    if len(aligned_ts_pres) > 1:
                        _, pressure_weights = self.calculate_total_volume(
                            aligned_flow_pres, aligned_ts_pres
                        )
                        inj_pres_avg = self.weighted_average_value_with_volume(
                            aligned_pres, aligned_ts_pres, pressure_weights
                        )

            # 4) inj_pres_max (maximum outlet pressure from injection pump)
            inj_pres_max = None
            if inj_pump_name is not None:
                self.select_unit(inj_pump_name)
                tagname_inj_pres_max = tag_lookup.get((inj_pump_name, "inj_pres_max"))
                values, _ = (
                    self.get_data(tagname_inj_pres_max) if tagname_inj_pres_max else (None, None)
                )
                if values is not None and len(values) > 0:
                    inj_pres_max = max(values)

            # 5) inj_inhibit_vol (total injected inhibitor volume - currently placeholder)
            self.select_unit(well_name)
            inj_inhibit_vol = None
            tag_inhibit = tag_lookup.get((well_name, "inj_inhibit_vol"))
            if tag_inhibit:
                values, datestamps = self.get_data(tag_inhibit)
                if values is not None and datestamps is not None:
                    inj_inhibit_vol, _ = self.calculate_total_volume(values, datestamps)

            # Collect row
            row_data = {
                "well_name": well_name,
                "inj_vol_water": inj_vol_water,
                "inj_temp_avg_weighted": inj_temp_avg_weighted,
                "inj_pres_avg": inj_pres_avg,
                "inj_pres_max": inj_pres_max,
                "inj_inhibit_vol": inj_inhibit_vol,
            }

            result_rows.append(row_data)

        # Build dataframe in correct column order
        df = pd.DataFrame(result_rows, columns=["well_name"] + param_order)
        for parameter in param_order:
            if parameter not in enabled_parameters:
                df[parameter] = None
        return df

    def calculate_nlog_2(self, rows_section2):
        """
        Calculate Section 2 NLOG values per doublet from flat rows_section2.

        rows_section2: list[dict] with keys:
            component_name, component_type, nlog_parameter, parameter, tagname, doublet

        Returns
        -------
        pd.DataFrame with columns:
            ["doublet", "tot_heat_MJ", "tot_oper_hours", "tot_el_cons_KWh"]
        """
        all_rows = rows_section2 or []
        enabled_nlog_parameters = {
            row.get("nlog_parameter") for row in all_rows if self._is_nlog_row_enabled(row)
        }
        rows_section2 = self._enabled_nlog_rows(all_rows)

        # Helpers --------------------------------------------------------------

        def normalize_doublet_name(doublet_raw: str) -> str:
            doublet_raw = (doublet_raw or "").strip()
            normalized = re.sub(r"\s+", "", doublet_raw.lower())
            return normalized or "default"

        def parse_utc_timestamp(timestamp_str: str) -> datetime:
            timestamp_str = timestamp_str.strip()
            if timestamp_str.endswith("Z"):
                timestamp_str = timestamp_str[:-1] + "+00:00"
            return datetime.fromisoformat(timestamp_str).astimezone(timezone.utc)

        def series_from_tagname(component_name: str, tagname: str) -> pd.Series:
            if not tagname:
                return pd.Series(dtype="float64")
            self.select_unit(component_name)
            values, datestamps = self.get_data(tagname)
            if values is None or datestamps is None:
                return pd.Series(dtype="float64")
            if len(values) == 0 or len(values) != len(datestamps):
                return pd.Series(dtype="float64")

            time_index = pd.to_datetime(
                [parse_utc_timestamp(ts) for ts in datestamps],
                utc=True,
            )
            series = pd.Series(list(values), index=time_index)
            series = pd.to_numeric(series, errors="coerce")
            series = series[~series.index.duplicated(keep="last")].sort_index()
            return series

        # Group rows by normalized doublet -------------------------------------

        doublet_groups = {}
        for row in rows_section2 or []:
            raw_doublet = row.get("doublet") or "default"
            norm_doublet = normalize_doublet_name(raw_doublet)
            if norm_doublet not in doublet_groups:
                doublet_groups[norm_doublet] = {
                    "label": raw_doublet,
                    "rows": [],
                }
            doublet_groups[norm_doublet]["rows"].append(row)

        result_rows = []

        # Process each doublet group -------------------------------------------

        for norm_doublet, group_data in doublet_groups.items():
            group_label = group_data["label"] or norm_doublet
            group_rows = group_data["rows"]

            # 1) Total heat (tot_heat_MJ) -------------------------------------
            total_heat_J = 0.0
            density_kg_per_m3 = 1000.0
            heat_capacity_J_per_kgK = 4186.0

            hex_rows = [
                row
                for row in group_rows
                if row.get("nlog_parameter") == "tot_heat_MJ"
                and row.get("component_type") == "heat_exchanger"
            ]

            hex_param_map = {}
            for row in hex_rows:
                hex_name = row["component_name"]
                parameter_name = row["parameter"]
                tagname = row["tagname"]
                hex_param_map.setdefault(hex_name, {})[parameter_name] = tagname

            for hex_name, param_tags in hex_param_map.items():
                tag_flow = param_tags.get("tot_heat_MJ_flow", "")
                tag_temp_inlet = param_tags.get("tot_heat_MJ_inlet_temp", "")
                tag_temp_outlet = param_tags.get("tot_heat_MJ_outlet_temp", "")

                series_flow = series_from_tagname(hex_name, tag_flow)
                series_temp_inlet = series_from_tagname(hex_name, tag_temp_inlet)
                series_temp_outlet = series_from_tagname(hex_name, tag_temp_outlet)

                if series_flow.empty or series_temp_inlet.empty or series_temp_outlet.empty:
                    continue

                union_index = (
                    series_flow.index.union(series_temp_inlet.index)
                    .union(series_temp_outlet.index)
                    .sort_values()
                )

                df_hex = pd.DataFrame(
                    {
                        "flow_m3h": series_flow.reindex(union_index),
                        "temp_in_C": series_temp_inlet.reindex(union_index),
                        "temp_out_C": series_temp_outlet.reindex(union_index),
                    },
                    index=union_index,
                ).interpolate(method="time", limit_direction="both")

                time_index = df_hex.index
                for interval_index in range(len(df_hex) - 1):
                    time_start = time_index[interval_index]
                    time_end = time_index[interval_index + 1]
                    delta_seconds = (time_end - time_start).total_seconds()
                    if delta_seconds <= 0:
                        continue

                    flow_start, flow_end = (
                        df_hex.iloc[interval_index]["flow_m3h"],
                        df_hex.iloc[interval_index + 1]["flow_m3h"],
                    )
                    temp_in_start, temp_in_end = (
                        df_hex.iloc[interval_index]["temp_in_C"],
                        df_hex.iloc[interval_index + 1]["temp_in_C"],
                    )
                    temp_out_start, temp_out_end = (
                        df_hex.iloc[interval_index]["temp_out_C"],
                        df_hex.iloc[interval_index + 1]["temp_out_C"],
                    )

                    if (
                        pd.isna(flow_start)
                        or pd.isna(flow_end)
                        or pd.isna(temp_in_start)
                        or pd.isna(temp_in_end)
                        or pd.isna(temp_out_start)
                        or pd.isna(temp_out_end)
                    ):
                        continue

                    flow_avg_m3h = 0.5 * (float(flow_start) + float(flow_end))
                    temp_in_avg = 0.5 * (float(temp_in_start) + float(temp_in_end))
                    temp_out_avg = 0.5 * (float(temp_out_start) + float(temp_out_end))

                    delta_temperature_K = temp_out_avg - temp_in_avg
                    mass_flow_kg_per_s = (flow_avg_m3h * density_kg_per_m3) / 3600.0
                    energy_J = (
                        mass_flow_kg_per_s
                        * heat_capacity_J_per_kgK
                        * delta_temperature_K
                        * delta_seconds
                    )
                    total_heat_J += energy_J

            total_heat_MJ = total_heat_J / 1_000_000.0

            # 2) Total operating hours (tot_oper_hours) ------------------------

            flow_threshold_m3_per_h = 1.0
            total_operating_hours = 0.0

            oper_rows = [row for row in group_rows if row.get("nlog_parameter") == "tot_oper_hours"]

            for row in oper_rows:
                well_name = row["component_name"]
                tag_flow = row["tagname"]

                series_flow = series_from_tagname(well_name, tag_flow)
                if series_flow.empty:
                    continue

                series_flow = series_flow.reindex(series_flow.index.sort_values())
                series_flow = series_flow.interpolate(method="time", limit_direction="both")

                time_index = series_flow.index
                for interval_index in range(len(series_flow) - 1):
                    time_start = time_index[interval_index]
                    time_end = time_index[interval_index + 1]
                    delta_seconds = (time_end - time_start).total_seconds()
                    if delta_seconds <= 0:
                        continue

                    flow_start = series_flow.iloc[interval_index]
                    flow_end = series_flow.iloc[interval_index + 1]
                    if pd.isna(flow_start) or pd.isna(flow_end):
                        continue

                    flow_avg = 0.5 * (float(flow_start) + float(flow_end))
                    if flow_avg > flow_threshold_m3_per_h:
                        total_operating_hours += delta_seconds / 3600.0

            # 3) Total electric consumption (tot_el_cons_KWh) ------------------

            electric_rows = [
                row
                for row in group_rows
                if row.get("nlog_parameter") == "tot_el_cons_KWh"
                and row.get("component_type") in ("esp", "injection_pump", "booster_pump")
            ]

            component_param_map = {}
            for row in electric_rows:
                component_name = row["component_name"]
                parameter_name = row["parameter"]
                tagname = row["tagname"]
                component_param_map.setdefault(component_name, {})[parameter_name] = tagname

            total_energy_Wh = 0.0
            power_factor = 1.0
            use_sqrt3 = False
            sqrt3_value = sqrt(3.0)

            for component_name, param_tags in component_param_map.items():
                tag_power = param_tags.get(
                    "tot_el_cons_KWh_power", ""
                )  # expected power_consumption tag (cumulative kWh)
                tag_current = param_tags.get("tot_el_cons_KWh_current", "")  # expected current tag
                tag_voltage = param_tags.get("tot_el_cons_KWh_voltage", "")  # expected voltage tag

                # Preferred: direct power consumption tag (cumulative energy meter in kWh)
                series_power = series_from_tagname(component_name, tag_power)
                if not series_power.empty:
                    series_power = series_power.reindex(series_power.index.sort_values())
                    series_power = series_power.interpolate(method="time", limit_direction="both")

                    # Power consumption data is cumulative energy (kWh), not instantaneous power
                    # Calculate final_reading_kWh - initial_reading_kWh,
                    # then convert to Wh for accumulation.
                    if len(series_power) >= 2:
                        energy_start_kWh = float(series_power.iloc[0])
                        energy_end_kWh = float(series_power.iloc[-1])
                        if not pd.isna(energy_start_kWh) and not pd.isna(energy_end_kWh):
                            total_energy_Wh += (energy_end_kWh - energy_start_kWh) * 1000.0

                    continue  # do not fall back if power series is available

                # Fallback: power = voltage * current
                series_current = series_from_tagname(component_name, tag_current)
                series_voltage = series_from_tagname(component_name, tag_voltage)
                if series_current.empty or series_voltage.empty:
                    continue

                union_index = series_current.index.union(series_voltage.index).sort_values()
                df_power = pd.DataFrame(
                    {
                        "current": series_current.reindex(union_index),
                        "voltage": series_voltage.reindex(union_index),
                    },
                    index=union_index,
                ).interpolate(method="time", limit_direction="both")

                time_index = df_power.index
                for interval_index in range(len(df_power) - 1):
                    time_start = time_index[interval_index]
                    time_end = time_index[interval_index + 1]
                    delta_seconds = (time_end - time_start).total_seconds()
                    if delta_seconds <= 0:
                        continue

                    current_start = df_power.iloc[interval_index]["current"]
                    current_end = df_power.iloc[interval_index + 1]["current"]
                    voltage_start = df_power.iloc[interval_index]["voltage"]
                    voltage_end = df_power.iloc[interval_index + 1]["voltage"]

                    if (
                        pd.isna(current_start)
                        or pd.isna(current_end)
                        or pd.isna(voltage_start)
                        or pd.isna(voltage_end)
                    ):
                        continue

                    current_avg = 0.5 * (float(current_start) + float(current_end))
                    voltage_avg = 0.5 * (float(voltage_start) + float(voltage_end))

                    power_W = voltage_avg * current_avg * float(power_factor)
                    if use_sqrt3:
                        power_W *= sqrt3_value

                    delta_hours = delta_seconds / 3600.0
                    total_energy_Wh += power_W * delta_hours

            total_energy_kWh = total_energy_Wh / 1000.0

            result_rows.append(
                {
                    "doublet": group_label,
                    "tot_heat_MJ": total_heat_MJ,
                    "tot_oper_hours": total_operating_hours,
                    "tot_el_cons_KWh": total_energy_kWh,
                }
            )

        section2_param_order = ["tot_heat_MJ", "tot_oper_hours", "tot_el_cons_KWh"]
        dataframe = pd.DataFrame(
            result_rows,
            columns=["doublet"] + section2_param_order,
        )
        for parameter in section2_param_order:
            if parameter not in enabled_nlog_parameters:
                dataframe[parameter] = None
        return dataframe

    def add_nlog_data(
        self,
        LicenseHolder,
        NlogPeriod,
        df_prod,
        df_inj,
        table3_df,
    ):
        """Load the NLOG EXCEL template and writes the data."""
        # Build template path
        if not hasattr(self, "project_path") or not self.project_path:
            raise ValueError("self.project_path is not set or is empty.")

        template_path = os.path.join(
            self.project_path,
            "_template",
            "report_generator",
            "aardwarmte_productiecijfers_template_v2025.xlsm",
        )

        if not os.path.exists(template_path):
            raise FileNotFoundError(f"NLOG template not found at: {template_path}")

        # Load workbook (keep formatting, keep_vba, keep formulas)
        wb = load_workbook(template_path, keep_vba=True, data_only=False)

        if "Aardwarmte" not in wb.sheetnames:
            raise ValueError('Sheet "Aardwarmte" not found in nlog_template.xlsm')

        ws = wb["Aardwarmte"]

        # --------------------------------------------------------------------------------------------
        # Write metadata (preserve formatting: only write values)
        # --------------------------------------------------------------------------------------------
        ws["B4"].value = LicenseHolder

        # NlogPeriod like "2025-05" -> "202505"
        nlog_period_compact = str(NlogPeriod).replace("-", "")
        ws["B5"].value = nlog_period_compact

        def _write_df_at(start_coord, df):
            if df is None or df.empty:
                return

            col_letter, start_row = coordinate_from_string(start_coord)
            start_col = column_index_from_string(col_letter)

            # Write values only (no headers), do not insert/delete rows/cols.
            for r_idx, row_vals in enumerate(df.itertuples(index=False, name=None), start=0):
                for c_idx, value in enumerate(row_vals, start=0):
                    ws.cell(row=start_row + r_idx, column=start_col + c_idx).value = value

        # --------------------------------------------------------------------------------------------
        # Write df_prod and df_inj with shared well_name column A and shading rules
        # --------------------------------------------------------------------------------------------
        from openpyxl.styles import PatternFill

        gray_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")

        start_row = 9

        # Column index helpers
        COL_A = 1
        COL_B = 2
        COL_J = column_index_from_string("J")  # prod data B-J
        COL_M = column_index_from_string("M")  # inj data M-Q
        COL_Q = column_index_from_string("Q")

        current_row = start_row

        # --- Production rows ---
        if df_prod is not None and not df_prod.empty:
            prod_cols = list(df_prod.columns)
            if "well_name" not in prod_cols:
                raise ValueError("df_prod must contain a 'well_name' column.")

            prod_other_cols = [c for c in prod_cols if c != "well_name"]

            for _, r in df_prod.iterrows():
                # Well name in column A
                ws.cell(row=current_row, column=COL_A).value = r["well_name"]

                # Remaining prod columns start at B (write across)
                for i, col_name in enumerate(prod_other_cols):
                    target_col = COL_B + i
                    if target_col > COL_J:
                        break
                    ws.cell(row=current_row, column=target_col).value = r[col_name]

                # Shade injection area (M-Q) gray for this production row
                for c in range(COL_M, COL_Q + 1):
                    ws.cell(row=current_row, column=c).fill = gray_fill

                current_row += 1

        # --- Injection rows ---
        if df_inj is not None and not df_inj.empty:
            inj_cols = list(df_inj.columns)
            if "well_name" not in inj_cols:
                raise ValueError("df_inj must contain a 'well_name' column.")

            inj_other_cols = [c for c in inj_cols if c != "well_name"]

            for _, r in df_inj.iterrows():
                # Well name in column A
                ws.cell(row=current_row, column=COL_A).value = r["well_name"]

                # Remaining inj columns start at M (write across)
                for i, col_name in enumerate(inj_other_cols):
                    target_col = COL_M + i
                    if target_col > COL_Q:
                        break
                    ws.cell(row=current_row, column=target_col).value = r[col_name]

                # Shade production area (B-J) gray for this injection row
                for c in range(COL_B, COL_J + 1):
                    ws.cell(row=current_row, column=c).fill = gray_fill

                current_row += 1

        # --------------------------------------------------------------------------------------------
        # Add Mining work table data
        # --------------------------------------------------------------------------------------------
        _write_df_at("A28", table3_df)

        return wb

    def get_unit_data(self, units, table_tagnames, use_plant_units_fallback=False):
        """Retrieve data for given units and tagnames from database."""
        table_data = {}
        for unit_name in units:
            self.select_unit(unit_name)
            table_data[unit_name] = {}

            for column_name in table_tagnames.keys():
                tagname = table_tagnames[column_name]["tagname"]
                unit_type = table_tagnames[column_name]["unit"]

                # Select correct unit
                self.select_unit(unit_name)
                if unit_type not in unit_name:
                    found = False

                    # First try: from connected units
                    for unit in self.unit.to_units:
                        if unit_type in unit.name:
                            self.select_unit(unit.name)
                            found = True
                            break

                    # Optional fallback: search through all plant units (used for injection)
                    if use_plant_units_fallback and not found:
                        for plant_unit in self.plant.units:
                            if unit_type in plant_unit.name:
                                self.select_unit(plant_unit.name)
                                break

                # Fetch data
                if tagname not in ["unknown", "loaded", "previous_tagname"]:
                    value_list, datestamp_list = self.get_data(tagname)
                    if not value_list:
                        table_data[unit_name][column_name] = {
                            "values": [],
                            "datestamps": [],
                            "status": "No data found",
                        }
                    else:
                        table_data[unit_name][column_name] = {
                            "values": value_list,
                            "datestamps": datestamp_list,
                            "status": "loaded",
                        }
                else:
                    table_data[unit_name][column_name] = {
                        "values": [],
                        "datestamps": [],
                        "status": tagname,
                    }

        return table_data

    # def calculate_total_heat_extracted_MJ(self, hex_data):
    #     """Calculate the total heat extracted from heat exchanger measured data."""
    #     rho_kg_m3 = 1000.0
    #     cp_J_kgK = 4186.0
    #
    #     def _parse_utc(ts: str) -> datetime:
    #         # Handles '...Z' and also already-offset strings.
    #         ts = ts.strip()
    #         if ts.endswith("Z"):
    #             ts = ts[:-1] + "+00:00"
    #         return datetime.fromisoformat(ts).astimezone(timezone.utc)
    #
    #     def _series_from(signal):
    #         vals = signal.get("values", []) or []
    #         dts = signal.get("datestamps", []) or []
    #         if len(vals) != len(dts) or len(vals) == 0:
    #             return pd.Series(dtype="float64")
    #
    #         idx = pd.to_datetime([_parse_utc(t) for t in dts], utc=True)
    #         s = pd.Series(list(vals), index=idx)
    #
    #         # Convert to numeric; non-numeric -> NaN
    #         s = pd.to_numeric(s, errors="coerce")
    #
    #         # If duplicates exist, keep the last reading per timestamp
    #         s = s[~s.index.duplicated(keep="last")].sort_index()
    #         return s
    #
    #     total_J = 0.0
    #
    #     for key, hx in (hex_data or {}).items():
    #         if "heat_exchanger" not in str(key):
    #             continue
    #         if not isinstance(hx, dict):
    #             continue
    #
    #         # If you want to enforce status, uncomment:
    #         # if hx.get("status") not in ("loaded", "ok", True):
    #         #     continue
    #
    #         flow_s = _series_from(hx.get("hex_secondary_flow", {}))
    #         tin_s = _series_from(hx.get("hex_secondary_inlet_temperature", {}))
    #         tout_s = _series_from(hx.get("hex_secondary_outlet_temperature", {}))
    #
    #         if flow_s.empty or tin_s.empty or tout_s.empty:
    #             continue
    #
    #         # Union of timestamps across signals
    #         union_index = flow_s.index.union(tin_s.index).union(tout_s.index).sort_values()
    #
    #         df = pd.DataFrame(
    #             {
    #                 "flow_m3h": flow_s.reindex(union_index),
    #                 "tin_C": tin_s.reindex(union_index),
    #                 "tout_C": tout_s.reindex(union_index),
    #             },
    #             index=union_index,
    #         )
    #
    #         # Interpolate in time (linear), consistent with "average between measured values"
    #         df = df.interpolate(method="time", limit_direction="both")
    #
    #         # Compute interval energy using average of endpoints (trapezoid)
    #         t = df.index
    #         for i in range(len(df) - 1):
    #             t0 = t[i]
    #             t1 = t[i + 1]
    #             dt_s = (t1 - t0).total_seconds()
    #             if dt_s <= 0:
    #                 continue
    #
    #             f0, f1 = df.iloc[i]["flow_m3h"], df.iloc[i + 1]["flow_m3h"]
    #             ti0, ti1 = df.iloc[i]["tin_C"], df.iloc[i + 1]["tin_C"]
    #             to0, to1 = df.iloc[i]["tout_C"], df.iloc[i + 1]["tout_C"]
    #
    #             # Skip intervals with NaN
    #             if (
    #                     pd.isna(f0)
    #                     or pd.isna(f1)
    #                     or pd.isna(ti0)
    #                     or pd.isna(ti1)
    #                     or pd.isna(to0)
    #                     or pd.isna(to1)
    #             ):
    #                 continue
    #
    #             flow_avg_m3h = 0.5 * (float(f0) + float(f1))
    #             tin_avg = 0.5 * (float(ti0) + float(ti1))
    #             tout_avg = 0.5 * (float(to0) + float(to1))
    #
    #             dT_K = tout_avg - tin_avg
    #             # If you want to ignore negative extraction, uncomment:
    #             # if dT_K <= 0:
    #             #     continue
    #
    #             m_dot_kg_s = (flow_avg_m3h * rho_kg_m3) / 3600.0
    #             Q_J = m_dot_kg_s * cp_J_kgK * dT_K * dt_s
    #             total_J += Q_J
    #
    #     return total_J / 1_000_000.0  # MJ

    # def calculate_esp_operational_hours_and_kwh(self, esp_data):
    #     """Calculate the total number of operational hours and the total power consumption."""
    #     current_threshold = 2.0
    #     power_factor = 1.0
    #     include_sqrt3 = False
    #
    #     def _parse_utc(ts):
    #         ts = ts.strip()
    #         if ts.endswith("Z"):
    #             ts = ts[:-1] + "+00:00"
    #         return datetime.fromisoformat(ts).astimezone(timezone.utc)
    #
    #     def _series_from(signal):
    #         vals = signal.get("values", []) or []
    #         dts = signal.get("datestamps", []) or []
    #         if len(vals) == 0 or len(vals) != len(dts):
    #             return pd.Series(dtype="float64")
    #
    #         idx = pd.to_datetime([_parse_utc(t) for t in dts], utc=True)
    #         s = pd.Series(list(vals), index=idx)
    #         s = pd.to_numeric(s, errors="coerce")
    #         s = s[~s.index.duplicated(keep="last")].sort_index()
    #         return s
    #
    #     if not esp_data:
    #         return 0.0, 0.0
    #
    #     esp_keys = [k for k in esp_data.keys() if str(k).startswith("esp_")]
    #     if not esp_keys:
    #         return 0.0, 0.0
    #
    #     total_running_hours_all_esps = 0.0
    #     total_energy_Wh_all_esps = 0.0
    #
    #     sqrt3 = sqrt(3.0)
    #
    #     for esp_key in esp_keys:
    #         esp = esp_data.get(esp_key, {})
    #         if not isinstance(esp, dict):
    #             continue
    #
    #         esp_current = _series_from(esp.get("esp_current", {}))
    #         esp_voltage = _series_from(esp.get("esp_voltage", {}))
    #
    #         if esp_current.empty or esp_voltage.empty:
    #             continue
    #
    #         # Align on union of timestamps and interpolate in time
    #         idx = esp_current.index.union(esp_voltage.index).sort_values()
    #         df = pd.DataFrame(
    #             {
    #                 "esp_current": esp_current.reindex(idx),
    #                 "esp_voltage": esp_voltage.reindex(idx),
    #             },
    #             index=idx,
    #         )
    #         df = df.interpolate(method="time", limit_direction="both")
    #
    #         # Integrate over intervals
    #         for i in range(len(df) - 1):
    #             t0, t1 = df.index[i], df.index[i + 1]
    #             dt_s = (t1 - t0).total_seconds()
    #             if dt_s <= 0:
    #                 continue
    #
    #             I0, I1 = df.iloc[i]["esp_current"], df.iloc[i + 1]["esp_current"]
    #             V0, V1 = df.iloc[i]["esp_voltage"], df.iloc[i + 1]["esp_voltage"]
    #             if pd.isna(I0) or pd.isna(I1) or pd.isna(V0) or pd.isna(V1):
    #                 continue
    #
    #             I_avg = 0.5 * (float(I0) + float(I1))
    #             V_avg = 0.5 * (float(V0) + float(V1))
    #
    #             dt_h = dt_s / 3600.0
    #
    #             # Running time rule
    #             if I_avg >= current_threshold:
    #                 total_running_hours_all_esps += dt_h
    #
    #             # Power + energy
    #             P_W = V_avg * I_avg * float(power_factor)
    #             if include_sqrt3:
    #                 P_W *= sqrt3
    #
    #             total_energy_Wh_all_esps += P_W * dt_h  # W * h = Wh
    #
    #     num_esps = len(esp_keys)
    #     operational_hours = total_running_hours_all_esps / num_esps if num_esps > 0 else 0.0
    #     electricity_consumption_kWh = total_energy_Wh_all_esps / 1000.0
    #
    #     return operational_hours, electricity_consumption_kWh

    def add_nlog_report(
        self, LicenseHolder, NlogPeriod, data_section1A, data_section1B, data_section2
    ):
        """Prepare NLOG report data and return Excel file as BytesIO."""
        # prod_table_data = self.get_unit_data(
        #     prod_wells, prod_table_tagnames, use_plant_units_fallback=False
        # )
        # inj_table_data = self.get_unit_data(
        #     inj_wells, inj_table_tagnames, use_plant_units_fallback=True
        # )
        # hex_data = self.get_unit_data(hexs, hex_tagnames, use_plant_units_fallback=True)
        # esp_data = self.get_unit_data(esps, esp_tagnames, use_plant_units_fallback=True)

        # ------------------------------------------------------------------------------------------------
        #                                   Prepare DataFrames
        # ------------------------------------------------------------------------------------------------
        # prod_table_rows = []
        # for well_name in prod_wells:
        #     row = {"well_name": well_name}

        #     # Water production volume
        #     if prod_table_data[well_name]["water_prod_volume"]["status"] == "loaded":
        #         total_volume, _ = self.calculate_total_volume(
        #             prod_table_data[well_name]["water_prod_volume"]["datestamps"],
        #             prod_table_data[well_name]["water_prod_volume"]["values"],
        #         )
        #         row["water_prod_volume"] = total_volume
        #     else:
        #         row["water_prod_volume"] = "No data found"
        #
        #     # Production pressure average
        #     if prod_table_data[well_name]["prod_pressure_avg"]["status"] == "loaded":
        #         _, volume_intervals = self.calculate_total_volume(
        #             prod_table_data[well_name]["water_prod_volume"]["datestamps"],
        #             prod_table_data[well_name]["water_prod_volume"]["values"],
        #         )
        #         row["prod_pressure_avg"] = self.weighted_average_value_with_volume(
        #             prod_table_data[well_name]["prod_pressure_avg"]["values"],
        #             prod_table_data[well_name]["prod_pressure_avg"]["datestamps"],
        #             volume_intervals,
        #         )
        #     else:
        #         row["prod_pressure_avg"] = "No data found"
        #
        #     # Production pressure min
        #     values = prod_table_data[well_name]["prod_pressure_avg"]["values"]
        #     row["prod_pressure_min"] = (
        #         min(v for v in values if v is not None) if values else "No data found"
        #     )
        #
        #     # Well pressure average
        #     if prod_table_data[well_name]["well_pressure_avg"]["status"] == "loaded":
        #         _, volume_intervals = self.calculate_total_volume(
        #             prod_table_data[well_name]["water_prod_volume"]["datestamps"],
        #             prod_table_data[well_name]["water_prod_volume"]["values"],
        #         )
        #         row["well_pressure_avg"] = self.weighted_average_value_with_volume(
        #             prod_table_data[well_name]["well_pressure_avg"]["values"],
        #             prod_table_data[well_name]["well_pressure_avg"]["datestamps"],
        #             volume_intervals,
        #         )
        #     else:
        #         row["well_pressure_avg"] = "No data found"
        #
        #     prod_table_rows.append(row)
        #
        # inj_table_rows = []
        # for well_name in inj_wells:
        #     row = {"well_name": well_name}
        #
        #     if inj_table_data[well_name]["water_inj_volume"]["status"] == "loaded":
        #         total_volume, _ = self.calculate_total_volume(
        #             inj_table_data[well_name]["water_inj_volume"]["datestamps"],
        #             inj_table_data[well_name]["water_inj_volume"]["values"],
        #         )
        #         row["water_inj_volume"] = total_volume
        #     else:
        #         row["water_inj_volume"] = "No data found"
        #
        #     if inj_table_data[well_name]["inj_temperature_avg"]["status"] == "loaded":
        #         _, volume_intervals = self.calculate_total_volume(
        #             inj_table_data[well_name]["water_inj_volume"]["datestamps"],
        #             inj_table_data[well_name]["water_inj_volume"]["values"],
        #         )
        #         row["inj_temperature_avg"] = self.weighted_average_value_with_volume(
        #             inj_table_data[well_name]["inj_temperature_avg"]["values"],
        #             inj_table_data[well_name]["inj_temperature_avg"]["datestamps"],
        #             volume_intervals,
        #         )
        #     else:
        #         row["inj_temperature_avg"] = "No data found"
        #
        #     if inj_table_data[well_name]["inj_pump_pressure_avg"]["status"] == "loaded":
        #         values = inj_table_data[well_name]["inj_pump_pressure_avg"]["values"]
        #         row["inj_pump_pressure_avg"] = sum(values) / len(values)
        #         row["inj_pump_pressure_max"] = max(v for v in values if v is not None)
        #     else:
        #         row["inj_pump_pressure_avg"] = "No data found"
        #         row["inj_pump_pressure_max"] = "No data found"
        #
        #     inj_table_rows.append(row)
        #
        # df_prod = pd.DataFrame(prod_table_rows)
        # df_inj = pd.DataFrame(inj_table_rows)
        #
        # # Create Mining work table dataframe
        # operational_hours, electricity_consumption_kWh = (
        #     self.calculate_esp_operational_hours_and_kwh(esp_data)
        # )
        # table3_dictionary = {
        #     "mining_work_tile": "Doublet 1",
        #     "total_extracted_heat": self.calculate_total_heat_extracted_MJ(hex_data),
        #     "operational_hours": operational_hours,
        #     "electricity_consumption_kWh": electricity_consumption_kWh,
        # }
        #
        # table3_df = pd.DataFrame([table3_dictionary])

        # ------------------------------------------------------------------------------------------------
        #                                   Load & modify Excel
        # ------------------------------------------------------------------------------------------------
        wb = self.add_nlog_data(
            LicenseHolder, NlogPeriod, data_section1A, data_section1B, data_section2
        )

        # ------------------------------------------------------------------------------------------------
        #                                   Return BytesIO for download
        # ------------------------------------------------------------------------------------------------
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def wrap_text(self, text, max_chars_per_line=90):
        """Change line in text output."""
        import textwrap

        return "\n".join(textwrap.wrap(text, max_chars_per_line))

    def add_text_section_page(self, user_text, section_title="User Input Section"):
        """Include the text section as a page."""
        # Check if user text is provided
        if not user_text or not user_text.strip():
            print("No user text input provided, skipping text section.")
            return

        # Prepare wrapped lines as a list of strings
        wrapped_lines = self.wrap_text(user_text, max_chars_per_line=90).split("\n")

        # A4 landscape size in inches
        # page_width, page_height = 11.69, 8.27

        # Parameters for text layout
        font_size = 12
        line_height = font_size * 1.2 / 72  # roughly converted from pts to inches (1pt=1/72 inch)
        top_margin = 0.8  # inches from top for title separation
        left_margin = 0.5
        # right_margin = 0.5
        bottom_margin = 0.5
        usable_height = self.page_size[0] - top_margin - bottom_margin

        # Number of lines that fit on one page
        max_lines_per_page = int(usable_height // line_height)

        # Split wrapped_lines into chunks fitting on pages
        chunks = [
            wrapped_lines[i : i + max_lines_per_page]
            for i in range(0, len(wrapped_lines), max_lines_per_page)
        ]

        for i, chunk in enumerate(chunks):
            fig, ax = plt.subplots(figsize=(self.page_size[0], self.page_size[1]))

            # Title only on first page or optionally on each page
            if i == 0:
                fig.suptitle(section_title, fontsize=16, fontweight="bold", y=0.95)

            ax.axis("off")

            # Join lines for this chunk and add text starting below title area
            # relative position for ax.text y coordinate
            y_start = 1 - top_margin / self.page_size[1]
            text = "\n".join(chunk)
            ax.text(
                left_margin / self.page_size[0],
                y_start,
                text,
                fontsize=font_size,
                va="top",
                ha="left",
                wrap=True,
                transform=ax.transAxes,
            )

            self.pdf_object.savefig(fig, bbox_inches="tight", pad_inches=0.3)
            plt.close(fig)

    def export_pdf(self):
        """Export PDF report."""
        self.pdf_object.close()
        print("Plots exported to plots.pdf")

        # Move to the beginning of the buffer
        self.pdf_buffer.seek(0)

    def load_section_config(self, folder: Path, json_name: str, default_name: str) -> dict:
        """Load section config JSON, falling back to default file if main file is missing."""
        json_path = folder / json_name
        default_path = folder / default_name

        if json_path.exists():
            return json.loads(json_path.read_text(encoding="utf-8"))

        if default_path.exists():
            data = json.loads(default_path.read_text(encoding="utf-8"))
            json_path.write_text(json.dumps(data, indent=2), encoding="utf-8")
            return data

        # Fallback: empty config if both files are missing
        json_path.write_text("{}", encoding="utf-8")
        return {}

    # def rows_from_section_config(self, section_cfg: dict, component_mapping: dict) -> list:
    #     """Build rows for all sections, handling nested section 2 structure properly."""
    #
    #     rows = []
    #
    #     for parameter, params_dict in section_cfg.items():
    #
    #         # -------------------------------
    #         # CASE 1: Flat structure (Section 1)
    #         # -------------------------------
    #         if "unit" in params_dict:
    #             component_type = params_dict["unit"]
    #
    #         else:
    #             # -----------------------------------------
    #             # CASE 2: Nested structure (Section 2)
    #             # Find the in-between key that contains a unit
    #             # -----------------------------------------
    #             component_type = ""
    #             for subkey, meta in params_dict.items():
    #                 if isinstance(meta, dict) and "unit" in meta:
    #                     component_type = meta["unit"]
    #                     break
    #
    #         # Skip if no component type identified
    #         unit_names = component_mapping.get(component_type, [])
    #         if not unit_names:
    #             continue
    #
    #         # -----------------------------------------
    #         # SECTION 1 FLAT FORMAT
    #         # -----------------------------------------
    #         if "tagname" in params_dict:
    #             tagname_value = params_dict.get("tagname", "")
    #             for unit_name in unit_names:
    #                 rows.append(
    #                     {
    #                         "component_name": unit_name,
    #                         "component_type": component_type,
    #                         "parameter": parameter,
    #                         "tagname": tagname_value,
    #                     }
    #                 )
    #             continue
    #
    #         # -----------------------------------------
    #         # SECTION 2 NESTED FORMAT
    #         # -----------------------------------------
    #         for subkey, meta in params_dict.items():
    #             if subkey == "unit":
    #                 continue
    #
    #             meta = meta or {}
    #             tagname_value = meta.get("tagname", "")
    #
    #             for unit_name in unit_names:
    #                 rows.append(
    #                     {
    #                         "component_name": unit_name,
    #                         "component_type": component_type,
    #                         "parameter": subkey,
    #                         "tagname": tagname_value,
    #                     }
    #                 )
    #
    #     return rows

    def rows_from_section1_config(self, section_cfg: dict, component_mapping: dict) -> list:
        """
        Build rows for Section 1A / Section 1B (flat structure).

        JSON format example::

            {
                "prod_pres_avg": { "tagname": "...", "unit": "production_well" }
            }
        """
        rows = []

        for parameter, meta in section_cfg.items():
            if not isinstance(meta, dict):
                continue

            # Component type comes directly from the 'unit' field
            component_type = meta.get("unit", "")
            unit_names = component_mapping.get(component_type, [])
            if not unit_names:
                continue

            tagname_value = meta.get("tagname", "")

            # Build a robust tagname that tolerates small naming inconsistencies
            # between component_type (may contain underscores) and tagname
            # prefixes (may omit underscores).
            full_tagname = self._compose_tagname(component_type, tagname_value)

            for unit_name in unit_names:
                rows.append(
                    {
                        "component_name": unit_name,
                        "component_type": component_type,
                        "parameter": parameter,
                        "tagname": full_tagname,
                        "enabled": True,
                    }
                )

        return rows

    def rows_from_section2_config(self, section_cfg: dict, component_mapping: dict) -> list:
        """
        Build rows for Section 2 (nested structure with in-between keys).

        Handles:
          - subkeys with a single unit:
                "unit": "esp"
          - subkeys with multiple units:
                "unit": ["esp", "booster_pump", "injection_pump"]

        Produces one row per:
            component_type × component_instance × subkey
        """
        rows = []

        for parameter, params_dict in section_cfg.items():
            if not isinstance(params_dict, dict):
                continue

            # ------------------------------------------------------------------
            # Determine all unit types referenced by this parameter
            # They may vary per subkey, so we will handle per-subkey logic below
            # ------------------------------------------------------------------

            for subkey, meta in params_dict.items():
                if subkey == "unit":
                    continue
                if not isinstance(meta, dict):
                    continue

                # Extract tagname and unit(s)
                tagname_value = meta.get("tagname", "")
                unit_field = meta.get("unit", "")

                # Standardize into list
                if isinstance(unit_field, str):
                    component_types = [unit_field]
                elif isinstance(unit_field, list):
                    component_types = unit_field
                else:
                    continue  # unsupported format

                # ------------------------------------------------------------------
                # Build rows for each component_type + component instance
                # ------------------------------------------------------------------
                for component_type in component_types:

                    # list of component names for this type
                    unit_names = component_mapping.get(component_type, [])
                    if not unit_names:
                        continue

                    for unit_name in unit_names:
                        # Build a robust tagname that tolerates small naming
                        # inconsistencies between component_type (may contain
                        # underscores) and tagname prefixes (may omit underscores).
                        full_tagname = self._compose_tagname(component_type, tagname_value)

                        rows.append(
                            {
                                "component_name": unit_name,
                                "component_type": component_type,
                                "nlog_parameter": parameter,
                                "parameter": subkey,  # subkey becomes parameter
                                "tagname": full_tagname,
                                "doublet": "Doublet 1",
                                "enabled": True,
                            }
                        )

        return rows

    def _load_or_create_section1_default_rows(
        self, default_file: Path, component_mapping: dict
    ) -> list:
        """Load Section 1 default rows, converting legacy default config files if needed."""
        default_data = json.loads(default_file.read_text(encoding="utf-8"))
        if isinstance(default_data, list):
            rows = self._ensure_nlog_enabled_default(default_data)
        elif isinstance(default_data, dict):
            rows = self.rows_from_section1_config(default_data, component_mapping)
        else:
            raise ValueError(f"Unsupported NLOG default file structure: {default_file}")

        default_file.write_text(json.dumps(rows, indent=2), encoding="utf-8")
        return rows

    def _load_or_create_section2_default_rows(
        self, default_file: Path, component_mapping: dict
    ) -> list:
        """Load Section 2 default rows, converting legacy default config files if needed."""
        default_data = json.loads(default_file.read_text(encoding="utf-8"))
        if isinstance(default_data, list):
            rows = self._ensure_nlog_enabled_default(default_data)
        elif isinstance(default_data, dict):
            rows = self.rows_from_section2_config(default_data, component_mapping)
        else:
            raise ValueError(f"Unsupported NLOG default file structure: {default_file}")

        default_file.write_text(json.dumps(rows, indent=2), encoding="utf-8")
        return rows

    def create_section1a_df(
        self, folder: Path, prod_wells: list, esps: list, aquifers: list
    ) -> pd.DataFrame:
        """Create dataframe for section 1A (production_well, esp, aquifer)."""
        folder.mkdir(parents=True, exist_ok=True)

        target_name = "tagnames_section1A.json"
        default_name = "tagnames_section1A_default.json"

        target_file = folder / target_name
        default_file = folder / default_name

        # -----------------------------------------------------------
        # CASE 1: Load existing section file
        # -----------------------------------------------------------
        if target_file.exists():
            rows = self._ensure_nlog_enabled_default(
                json.loads(target_file.read_text(encoding="utf-8"))
            )
            return pd.DataFrame(
                rows,
                columns=["component_name", "component_type", "parameter", "tagname", "enabled"],
            )

        # -----------------------------------------------------------
        # CASE 2: Create section from default file
        # -----------------------------------------------------------
        if not default_file.exists():
            raise FileNotFoundError(f"Default file missing: {default_file}")

        component_mapping = {
            "production_well": prod_wells,
            "esp": esps,
            "aquifer": aquifers,
        }

        rows = self._load_or_create_section1_default_rows(default_file, component_mapping)

        # Save default rows to the user-editable settings file
        target_file.write_text(json.dumps(rows, indent=2), encoding="utf-8")

        return pd.DataFrame(
            rows, columns=["component_name", "component_type", "parameter", "tagname", "enabled"]
        )

    def create_section1b_df(
        self, folder: Path, inj_wells: list, hexs: list, injection_pumps: list
    ) -> pd.DataFrame:
        """Create dataframe for section 1B (injection_well, heat_exchanger, injection_pump)."""
        folder.mkdir(parents=True, exist_ok=True)

        target_name = "tagnames_section1B.json"
        default_name = "tagnames_section1B_default.json"

        target_file = folder / target_name
        default_file = folder / default_name

        # ---------- Case 1: file already exists -> load rows ----------
        if target_file.exists():
            rows = self._ensure_nlog_enabled_default(
                json.loads(target_file.read_text(encoding="utf-8"))
            )
            return pd.DataFrame(
                rows,
                columns=["component_name", "component_type", "parameter", "tagname", "enabled"],
            )

        # ---------- Case 2: file missing -> create from default ----------
        if not default_file.exists():
            raise FileNotFoundError(f"Default file missing: {default_file}")

        component_mapping = {
            "injection_well": inj_wells,
            "heat_exchanger": hexs,
            "injection_pump": injection_pumps,
        }

        rows = self._load_or_create_section1_default_rows(default_file, component_mapping)

        # Save default rows to the user-editable settings file
        target_file.write_text(json.dumps(rows, indent=2), encoding="utf-8")

        return pd.DataFrame(
            rows,
            columns=["component_name", "component_type", "parameter", "tagname", "enabled"],
        )

    def create_section2_df(
        self,
        folder: Path,
        prod_wells: list,
        hexs: list,
        esps: list,
        booster_pumps: list,
        injection_pumps: list,
    ) -> pd.DataFrame:
        """Create dataframe for section 2.

        Covers production wells, heat exchangers, ESPs, booster pumps,
        and injection pumps.
        """
        folder.mkdir(parents=True, exist_ok=True)

        target_name = "tagnames_section2.json"
        default_name = "tagnames_section2_default.json"

        target_file = folder / target_name
        default_file = folder / default_name

        # ---------- Case 1: file already exists -> load rows ----------
        if target_file.exists():
            rows = self._ensure_nlog_enabled_default(
                json.loads(target_file.read_text(encoding="utf-8"))
            )
            return pd.DataFrame(
                rows,
                columns=[
                    "component_name",
                    "component_type",
                    "nlog_parameter",
                    "parameter",
                    "tagname",
                    "doublet",
                    "enabled",
                ],
            )

        # ---------- Case 2: file missing -> create from default ----------
        if not default_file.exists():
            raise FileNotFoundError(f"Default file missing: {default_file}")

        component_mapping = {
            "production_well": prod_wells,
            "heat_exchanger": hexs,
            "esp": esps,
            "booster_pump": booster_pumps,
            "injection_pump": injection_pumps,
        }

        rows = self._load_or_create_section2_default_rows(default_file, component_mapping)

        # Save default rows to the user-editable settings file
        target_file.write_text(json.dumps(rows, indent=2), encoding="utf-8")

        return pd.DataFrame(
            rows,
            columns=[
                "component_name",
                "component_type",
                "nlog_parameter",
                "parameter",
                "tagname",
                "doublet",
                "enabled",
            ],
        )

    def get_nlog_tagnames_df(
        self,
        folder: Path,
        inj_wells: list,
        prod_wells: list,
        esps: list,
        hexs: list,
        booster_pumps: list,
        injection_pumps: list,
        aquifers: list,
    ):
        """Build dataframes for sections 1A, 1B, and 2."""
        section1a_df = self.create_section1a_df(
            folder=folder,
            prod_wells=prod_wells,
            esps=esps,
            aquifers=aquifers,
        )

        section1b_df = self.create_section1b_df(
            folder=folder,
            inj_wells=inj_wells,
            hexs=hexs,
            injection_pumps=injection_pumps,
        )

        section2_df = self.create_section2_df(
            folder=folder,
            prod_wells=prod_wells,
            hexs=hexs,
            esps=esps,
            booster_pumps=booster_pumps,
            injection_pumps=injection_pumps,
        )

        return section1a_df, section1b_df, section2_df
